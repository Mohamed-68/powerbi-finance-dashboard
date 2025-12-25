"""
Generate a realistic 5-year (2020–2024) finance dataset for a Power BI portfolio project.

What this simulates (aligned with the job description):
- Trial balance / P&L style financial data (Actual + Budget)
- Operational data (Sales + Headcount)
- Messy Excel P&L statement sheets to practice Power Query clean + unpivot
- Clean facts + dimensions to practice best-practice data modeling in Power BI

Outputs:
- data/Financial_Raw_Data_2020_2024.xlsx

Run:
  python scripts/generate_dataset.py

Requires:
  pip install -r requirements.txt
"""

from __future__ import annotations

import math
import random
from dataclasses import dataclass
from typing import Dict, List, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# -----------------------------
# CONFIG
# -----------------------------

SEED = 42
random.seed(SEED)
np.random.seed(SEED)

YEARS = list(range(2020, 2025))  # last 5 years: 2020–2024
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# month seasonality (normalized)
SEASONALITY = np.array([0.92, 0.95, 1.02, 1.00, 1.03, 1.05, 1.04, 1.02, 0.98, 1.03, 1.06, 1.10])
SEASONALITY = SEASONALITY / SEASONALITY.mean()

CURRENCY = "EUR"
COMPANY = "Company XYZ"


# -----------------------------
# DIMENSIONS
# -----------------------------

ENTITIES = [
    {"EntityKey": 1, "Entity": "France", "Region": "EU-West"},
    {"EntityKey": 2, "Entity": "Germany", "Region": "EU-Central"},
    {"EntityKey": 3, "Entity": "Spain", "Region": "EU-South"},
]

COST_CENTERS = [
    {"CostCenterKey": 10, "CostCenter": "Sales"},
    {"CostCenterKey": 20, "CostCenter": "Operations"},
    {"CostCenterKey": 30, "CostCenter": "G&A"},
]

PRODUCTS = [
    {"ProductKey": 100, "Product": "Core", "BasePrice": 120.0, "COGSRate": 0.44},
    {"ProductKey": 200, "Product": "Pro", "BasePrice": 210.0, "COGSRate": 0.48},
    {"ProductKey": 300, "Product": "Services", "BasePrice": 95.0, "COGSRate": 0.28},
]


@dataclass
class Account:
    AccountCode: str
    AccountName: str
    PnLGroup: str      # Revenue / COGS / Opex / Other
    Level1: str        # P&L major section
    Level2: str        # Subsection
    SortOrder: int
    NaturalSign: int   # +1 revenue/income, -1 expense


def build_accounts() -> List[Account]:
    accounts: List[Account] = []

    # Revenue (by product)
    accounts += [
        Account("4001", "Revenue - Core", "Revenue", "Revenue", "Product Revenue", 110, +1),
        Account("4002", "Revenue - Pro", "Revenue", "Revenue", "Product Revenue", 120, +1),
        Account("4010", "Revenue - Services", "Revenue", "Revenue", "Service Revenue", 130, +1),
    ]

    # COGS (by type)
    accounts += [
        Account("5001", "COGS - Materials", "COGS", "COGS", "Direct Costs", 210, -1),
        Account("5002", "COGS - Direct Labor", "COGS", "COGS", "Direct Costs", 220, -1),
        Account("5003", "COGS - Freight & Logistics", "COGS", "COGS", "Direct Costs", 230, -1),
    ]

    # Opex
    accounts += [
        Account("6001", "Opex - Salaries & Wages", "Opex", "Operating Expenses", "People", 310, -1),
        Account("6002", "Opex - Payroll Taxes", "Opex", "Operating Expenses", "People", 320, -1),
        Account("6003", "Opex - Rent", "Opex", "Operating Expenses", "Facilities", 330, -1),
        Account("6004", "Opex - Marketing", "Opex", "Operating Expenses", "Commercial", 340, -1),
        Account("6005", "Opex - Software & Subscriptions", "Opex", "Operating Expenses", "IT", 350, -1),
        Account("6006", "Opex - Travel", "Opex", "Operating Expenses", "Commercial", 360, -1),
    ]

    # Other
    accounts += [
        Account("7001", "Other - Interest Income", "Other", "Other", "Finance", 410, +1),
        Account("7002", "Other - Interest Expense", "Other", "Other", "Finance", 420, -1),
    ]

    return accounts


# -----------------------------
# DATE DIM
# -----------------------------

def build_dim_date(years: List[int]) -> pd.DataFrame:
    rows = []
    for y in years:
        for m in range(1, 13):
            month_end = pd.Timestamp(y, m, 1) + pd.offsets.MonthEnd(0)
            rows.append(
                {
                    "DateKey": int(month_end.strftime("%Y%m%d")),
                    "MonthEndDate": month_end.date(),
                    "Year": y,
                    "MonthNumber": m,
                    "MonthName": MONTHS[m - 1],
                    "Quarter": f"Q{((m - 1) // 3) + 1}",
                    "MonthStartDate": (pd.Timestamp(y, m, 1)).date(),
                }
            )
    return pd.DataFrame(rows)


# -----------------------------
# OPERATIONAL DATA GENERATION
# -----------------------------

def year_factor_actual(year: int) -> float:
    base = {2020: 0.92, 2021: 1.05, 2022: 1.10, 2023: 1.14, 2024: 1.20}
    return base.get(year, 1.0)


def year_factor_budget(year: int) -> float:
    base = {2020: 1.02, 2021: 1.06, 2022: 1.12, 2023: 1.16, 2024: 1.22}
    return base.get(year, 1.0)


def inflation_factor(year: int) -> float:
    base = {2020: 1.00, 2021: 1.02, 2022: 1.08, 2023: 1.12, 2024: 1.15}
    return base.get(year, 1.0)


def pandemic_q2_shock(year: int, month: int) -> float:
    if year == 2020 and month in (4, 5, 6):
        return 0.62
    return 1.0


def make_sales_operational(dim_date: pd.DataFrame) -> pd.DataFrame:
    rng = np.random.default_rng(SEED + 10)
    rows = []

    base_units = {
        ("France", "Core"): 5200,
        ("France", "Pro"): 2400,
        ("France", "Services"): 3300,
        ("Germany", "Core"): 6100,
        ("Germany", "Pro"): 2600,
        ("Germany", "Services"): 2900,
        ("Spain", "Core"): 4200,
        ("Spain", "Pro"): 1800,
        ("Spain", "Services"): 2600,
    }

    prod_lookup = {p["Product"]: p for p in PRODUCTS}

    for _, d in dim_date.iterrows():
        y = int(d["Year"])
        m = int(d["MonthNumber"])
        season = float(SEASONALITY[m - 1])

        for ent in [e["Entity"] for e in ENTITIES]:
            for prod in prod_lookup.keys():
                base = base_units[(ent, prod)]
                units = base * year_factor_actual(y) * season * pandemic_q2_shock(y, m)
                units *= (1.0 + rng.normal(0, 0.05))
                units = max(units, 0)

                price = prod_lookup[prod]["BasePrice"] * inflation_factor(y) * (1.0 + rng.normal(0, 0.02))
                price = max(price, 1)

                revenue = units * price

                rows.append(
                    {
                        "MonthEndDate": d["MonthEndDate"],
                        "Year": y,
                        "MonthNumber": m,
                        "Entity": ent,
                        "Product": prod,
                        "Units": float(round(units, 2)),
                        "AvgPrice": float(round(price, 2)),
                        "Revenue": float(round(revenue, 2)),
                    }
                )

    return pd.DataFrame(rows)


def make_headcount_operational(dim_date: pd.DataFrame) -> pd.DataFrame:
    rng = np.random.default_rng(SEED + 20)
    rows = []

    base_hc = {
        ("France", "Sales"): 18,
        ("France", "Operations"): 26,
        ("France", "G&A"): 12,
        ("Germany", "Sales"): 20,
        ("Germany", "Operations"): 30,
        ("Germany", "G&A"): 14,
        ("Spain", "Sales"): 14,
        ("Spain", "Operations"): 22,
        ("Spain", "G&A"): 10,
    }

    base_salary = {"Sales": 5200, "Operations": 4700, "G&A": 5600}

    for _, d in dim_date.iterrows():
        y = int(d["Year"])
        m = int(d["MonthNumber"])

        growth = {2020: 0.98, 2021: 1.01, 2022: 1.03, 2023: 1.04, 2024: 1.05}.get(y, 1.0)

        for ent in [e["Entity"] for e in ENTITIES]:
            for cc in [c["CostCenter"] for c in COST_CENTERS]:
                hc = base_hc[(ent, cc)] * growth
                hc *= (1.0 + rng.normal(0, 0.02))
                hc = max(hc, 0)

                salary = base_salary[cc] * inflation_factor(y) * (1.0 + rng.normal(0, 0.015))
                salary = max(salary, 1)

                rows.append(
                    {
                        "MonthEndDate": d["MonthEndDate"],
                        "Year": y,
                        "MonthNumber": m,
                        "Entity": ent,
                        "CostCenter": cc,
                        "Headcount": float(round(hc, 2)),
                        "AvgMonthlySalary": float(round(salary, 2)),
                    }
                )

    return pd.DataFrame(rows)


# -----------------------------
# FACT P&L (ACTUAL + BUDGET)
# -----------------------------

def build_fact_pnl(
    dim_date: pd.DataFrame,
    accounts: List[Account],
    sales_ops: pd.DataFrame,
    headcount_ops: pd.DataFrame,
) -> pd.DataFrame:
    rng = np.random.default_rng(SEED + 30)
    acc = pd.DataFrame([a.__dict__ for a in accounts])

    prod_rates = {p["Product"]: p["COGSRate"] for p in PRODUCTS}

    rows = []

    sales_key = ["MonthEndDate", "Entity", "Product"]
    sales = sales_ops.groupby(sales_key, as_index=False).agg({"Revenue": "sum"})

    sales_budget = sales.copy()
    sales_budget["Revenue"] = sales_budget.apply(
        lambda r: r["Revenue"]
        * (year_factor_budget(int(str(r["MonthEndDate"])[:4])) / year_factor_actual(int(str(r["MonthEndDate"])[:4])))
        * (1.0 + rng.normal(0, 0.02)),
        axis=1,
    )
    sales_budget["Revenue"] = sales_budget["Revenue"].clip(lower=0)

    hc_key = ["MonthEndDate", "Entity", "CostCenter"]
    hc = headcount_ops.groupby(hc_key, as_index=False).agg({"Headcount": "mean", "AvgMonthlySalary": "mean"})

    hc_budget = hc.copy()
    hc_budget["Headcount"] = hc_budget["Headcount"] * (1.0 + rng.normal(0.01, 0.01, size=len(hc_budget)))
    hc_budget["AvgMonthlySalary"] = hc_budget["AvgMonthlySalary"] * (1.0 + rng.normal(0.01, 0.01, size=len(hc_budget)))

    def fixed_rent(entity: str) -> float:
        base = {"France": 18_000, "Germany": 22_000, "Spain": 14_000}
        return base[entity]

    def fixed_software(entity: str) -> float:
        base = {"France": 9_000, "Germany": 10_500, "Spain": 7_000}
        return base[entity]

    for _, d in dim_date.iterrows():
        date = d["MonthEndDate"]
        y = int(d["Year"])
        m = int(d["MonthNumber"])

        for ent in [e["Entity"] for e in ENTITIES]:

            for prod in [p["Product"] for p in PRODUCTS]:
                rev_actual = float(
                    sales.loc[
                        (sales["MonthEndDate"] == date) & (sales["Entity"] == ent) & (sales["Product"] == prod),
                        "Revenue",
                    ].values[0]
                )
                rev_budget = float(
                    sales_budget.loc[
                        (sales_budget["MonthEndDate"] == date)
                        & (sales_budget["Entity"] == ent)
                        & (sales_budget["Product"] == prod),
                        "Revenue",
                    ].values[0]
                )

                rev_account = {"Core": "4001", "Pro": "4002", "Services": "4010"}[prod]

                rows.append(
                    {"MonthEndDate": date, "Scenario": "Actual", "Entity": ent, "CostCenter": "Sales", "Product": prod, "AccountCode": rev_account, "Amount": rev_actual}
                )
                rows.append(
                    {"MonthEndDate": date, "Scenario": "Budget", "Entity": ent, "CostCenter": "Sales", "Product": prod, "AccountCode": rev_account, "Amount": rev_budget}
                )

                cogs_rate = prod_rates[prod]

                materials = -(rev_actual * (cogs_rate * 0.62)) * (1.0 + rng.normal(0, 0.03))
                labor = -(rev_actual * (cogs_rate * 0.28)) * (1.0 + rng.normal(0, 0.02))
                freight = -(rev_actual * (cogs_rate * 0.10)) * (1.0 + rng.normal(0, 0.06))

                if y in (2022, 2023):
                    materials *= 1.06
                    freight *= 1.08

                materials_b = -(rev_budget * (cogs_rate * 0.62)) * (1.0 + rng.normal(0, 0.01))
                labor_b = -(rev_budget * (cogs_rate * 0.28)) * (1.0 + rng.normal(0, 0.01))
                freight_b = -(rev_budget * (cogs_rate * 0.10)) * (1.0 + rng.normal(0, 0.02))

                rows.append({"MonthEndDate": date, "Scenario": "Actual", "Entity": ent, "CostCenter": "Operations", "Product": prod, "AccountCode": "5001", "Amount": float(materials)})
                rows.append({"MonthEndDate": date, "Scenario": "Actual", "Entity": ent, "CostCenter": "Operations", "Product": prod, "AccountCode": "5002", "Amount": float(labor)})
                rows.append({"MonthEndDate": date, "Scenario": "Actual", "Entity": ent, "CostCenter": "Operations", "Product": prod, "AccountCode": "5003", "Amount": float(freight)})

                rows.append({"MonthEndDate": date, "Scenario": "Budget", "Entity": ent, "CostCenter": "Operations", "Product": prod, "AccountCode": "5001", "Amount": float(materials_b)})
                rows.append({"MonthEndDate": date, "Scenario": "Budget", "Entity": ent, "CostCenter": "Operations", "Product": prod, "AccountCode": "5002", "Amount": float(labor_b)})
                rows.append({"MonthEndDate": date, "Scenario": "Budget", "Entity": ent, "CostCenter": "Operations", "Product": prod, "AccountCode": "5003", "Amount": float(freight_b)})

            rev_a_total = sales.loc[(sales["MonthEndDate"] == date) & (sales["Entity"] == ent), "Revenue"].sum()
            rev_b_total = sales_budget.loc[(sales_budget["MonthEndDate"] == date) & (sales_budget["Entity"] == ent), "Revenue"].sum()

            for cc in [c["CostCenter"] for c in COST_CENTERS]:
                hca = hc.loc[(hc["MonthEndDate"] == date) & (hc["Entity"] == ent) & (hc["CostCenter"] == cc)]
                hcb = hc_budget.loc[(hc_budget["MonthEndDate"] == date) & (hc_budget["Entity"] == ent) & (hc_budget["CostCenter"] == cc)]

                head_a = float(hca["Headcount"].values[0])
                sal_a = float(hca["AvgMonthlySalary"].values[0])
                head_b = float(hcb["Headcount"].values[0])
                sal_b = float(hcb["AvgMonthlySalary"].values[0])

                salaries_a = -(head_a * sal_a) * (1.0 + rng.normal(0, 0.01))
                taxes_a = salaries_a * 0.14
                salaries_b = -(head_b * sal_b) * (1.0 + rng.normal(0, 0.005))
                taxes_b = salaries_b * 0.14

                rows.append({"MonthEndDate": date, "Scenario": "Actual", "Entity": ent, "CostCenter": cc, "Product": "", "AccountCode": "6001", "Amount": float(salaries_a)})
                rows.append({"MonthEndDate": date, "Scenario": "Actual", "Entity": ent, "CostCenter": cc, "Product": "", "AccountCode": "6002", "Amount": float(taxes_a)})
                rows.append({"MonthEndDate": date, "Scenario": "Budget", "Entity": ent, "CostCenter": cc, "Product": "", "AccountCode": "6001", "Amount": float(salaries_b)})
                rows.append({"MonthEndDate": date, "Scenario": "Budget", "Entity": ent, "CostCenter": cc, "Product": "", "AccountCode": "6002", "Amount": float(taxes_b)})

            rent_a = -(fixed_rent(ent) * inflation_factor(y)) * (1.0 + rng.normal(0, 0.01))
            soft_a = -(fixed_software(ent) * inflation_factor(y)) * (1.0 + rng.normal(0, 0.02))
            rent_b = -(fixed_rent(ent) * inflation_factor(y)) * (1.0 + rng.normal(0, 0.005))
            soft_b = -(fixed_software(ent) * inflation_factor(y)) * (1.0 + rng.normal(0, 0.01))

            rows.append({"MonthEndDate": date, "Scenario": "Actual", "Entity": ent, "CostCenter": "G&A", "Product": "", "AccountCode": "6003", "Amount": float(rent_a)})
            rows.append({"MonthEndDate": date, "Scenario": "Actual", "Entity": ent, "CostCenter": "G&A", "Product": "", "AccountCode": "6005", "Amount": float(soft_a)})
            rows.append({"MonthEndDate": date, "Scenario": "Budget", "Entity": ent, "CostCenter": "G&A", "Product": "", "AccountCode": "6003", "Amount": float(rent_b)})
            rows.append({"MonthEndDate": date, "Scenario": "Budget", "Entity": ent, "CostCenter": "G&A", "Product": "", "AccountCode": "6005", "Amount": float(soft_b)})

            mkt_rate_a = 0.065 if y >= 2021 else 0.055
            mkt_rate_b = 0.072
            marketing_a = -(rev_a_total * mkt_rate_a) * (1.0 + rng.normal(0, 0.07))
            marketing_b = -(rev_b_total * mkt_rate_b) * (1.0 + rng.normal(0, 0.03))

            travel_base = 0.010 if y == 2020 else (0.015 if y == 2021 else 0.020)
            travel_a = -(rev_a_total * travel_base) * pandemic_q2_shock(y, m) * (1.0 + rng.normal(0, 0.12))
            travel_b = -(rev_b_total * 0.020) * (1.0 + rng.normal(0, 0.05))

            rows.append({"MonthEndDate": date, "Scenario": "Actual", "Entity": ent, "CostCenter": "Sales", "Product": "", "AccountCode": "6004", "Amount": float(marketing_a)})
            rows.append({"MonthEndDate": date, "Scenario": "Budget", "Entity": ent, "CostCenter": "Sales", "Product": "", "AccountCode": "6004", "Amount": float(marketing_b)})
            rows.append({"MonthEndDate": date, "Scenario": "Actual", "Entity": ent, "CostCenter": "Sales", "Product": "", "AccountCode": "6006", "Amount": float(travel_a)})
            rows.append({"MonthEndDate": date, "Scenario": "Budget", "Entity": ent, "CostCenter": "Sales", "Product": "", "AccountCode": "6006", "Amount": float(travel_b)})

            interest_income_a = (950 + 90 * math.sin(m / 12 * 2 * math.pi)) * (1.0 + rng.normal(0, 0.20))
            interest_income_b = (1050 + 60 * math.sin(m / 12 * 2 * math.pi)) * (1.0 + rng.normal(0, 0.10))

            debt_proxy = max(0, (-(rent_a + soft_a)) * 0.6)
            interest_exp_a = -(debt_proxy * 0.08) * (1.0 + rng.normal(0, 0.15))
            interest_exp_b = -(debt_proxy * 0.075) * (1.0 + rng.normal(0, 0.08))

            rows.append({"MonthEndDate": date, "Scenario": "Actual", "Entity": ent, "CostCenter": "G&A", "Product": "", "AccountCode": "7001", "Amount": float(interest_income_a)})
            rows.append({"MonthEndDate": date, "Scenario": "Budget", "Entity": ent, "CostCenter": "G&A", "Product": "", "AccountCode": "7001", "Amount": float(interest_income_b)})
            rows.append({"MonthEndDate": date, "Scenario": "Actual", "Entity": ent, "CostCenter": "G&A", "Product": "", "AccountCode": "7002", "Amount": float(interest_exp_a)})
            rows.append({"MonthEndDate": date, "Scenario": "Budget", "Entity": ent, "CostCenter": "G&A", "Product": "", "AccountCode": "7002", "Amount": float(interest_exp_b)})

    fact = pd.DataFrame(rows)

    fact = fact.merge(acc, on="AccountCode", how="left")
    fact = fact.merge(dim_date[["MonthEndDate", "Year", "MonthNumber", "Quarter"]], on="MonthEndDate", how="left")
    fact["Amount"] = fact["Amount"].astype(float).round(2)

    cols = [
        "MonthEndDate", "Year", "Quarter", "MonthNumber",
        "Scenario",
        "Entity", "CostCenter", "Product",
        "AccountCode", "AccountName",
        "PnLGroup", "Level1", "Level2", "SortOrder",
        "Amount",
    ]
    return fact[cols]


def build_trial_balance(fact_pnl: pd.DataFrame) -> pd.DataFrame:
    tb = (
        fact_pnl.groupby(
            ["MonthEndDate", "Year", "MonthNumber", "Quarter", "Scenario", "Entity", "CostCenter", "AccountCode", "AccountName"],
            as_index=False
        )
        .agg(NetAmount=("Amount", "sum"))
    )

    tb["Debit"] = tb["NetAmount"].apply(lambda x: round(-x, 2) if x < 0 else 0.0)
    tb["Credit"] = tb["NetAmount"].apply(lambda x: round(x, 2) if x > 0 else 0.0)
    tb["SourceSystem"] = np.where(tb["Scenario"] == "Actual", "ERP_GL", "FP&A_BudgetTool")

    cols = [
        "MonthEndDate", "Year", "Quarter", "MonthNumber",
        "Scenario", "SourceSystem",
        "Entity", "CostCenter",
        "AccountCode", "AccountName",
        "Debit", "Credit", "NetAmount",
    ]
    return tb[cols]


# -----------------------------
# "MESSY" P&L STATEMENTS (WIDE FORMAT)
# -----------------------------

def to_pnl_lines_consolidated(fact_pnl: pd.DataFrame, year: int, scenario: str) -> pd.DataFrame:
    f = fact_pnl[(fact_pnl["Year"] == year) & (fact_pnl["Scenario"] == scenario)].copy()

    f = (
        f.groupby(["AccountCode", "AccountName", "PnLGroup", "Level1", "Level2", "MonthNumber"], as_index=False)
         .agg(Amount=("Amount", "sum"))
    )

    pivot = f.pivot_table(
        index=["AccountCode", "AccountName", "PnLGroup", "Level1", "Level2"],
        columns="MonthNumber",
        values="Amount",
        aggfunc="sum",
        fill_value=0.0,
    ).reset_index()

    def month_vec(df_subset: pd.DataFrame) -> np.ndarray:
        if df_subset.empty:
            return np.zeros(12)
        vals = []
        for m in range(1, 13):
            vals.append(float(df_subset.get(m, 0.0).sum()) if m in df_subset.columns else 0.0)
        return np.array(vals, dtype=float)

    rev = pivot[pivot["PnLGroup"] == "Revenue"]
    cogs = pivot[pivot["PnLGroup"] == "COGS"]
    opex = pivot[pivot["PnLGroup"] == "Opex"]
    other = pivot[pivot["PnLGroup"] == "Other"]

    total_revenue = month_vec(rev)
    total_cogs = month_vec(cogs)
    gross_profit = total_revenue + total_cogs
    total_opex = month_vec(opex)
    ebitda = gross_profit + total_opex
    total_other = month_vec(other)
    net_income = ebitda + total_other

    rows = []

    def add_section(title: str):
        rows.append({"Account": "", "Line Item": title, **{m: "" for m in MONTHS}, "Total": ""})

    def add_blank():
        rows.append({"Account": "", "Line Item": "", **{m: "" for m in MONTHS}, "Total": ""})

    def add_line(acc_code: str, name: str, vec: np.ndarray):
        r = {"Account": acc_code, "Line Item": name}
        for i, mon in enumerate(MONTHS):
            r[mon] = float(vec[i])
        r["Total"] = float(vec.sum())
        rows.append(r)

    add_section("REVENUE")
    for _, row in rev.sort_values(["AccountCode"]).iterrows():
        vec = np.array([float(row.get(m, 0.0)) for m in range(1, 13)], dtype=float)
        add_line(str(row["AccountCode"]), "  " + str(row["AccountName"]).replace("Revenue - ", ""), vec)
    add_line("", "Total Revenue", total_revenue)

    add_blank()
    add_section("COST OF GOODS SOLD (COGS)")
    for _, row in cogs.sort_values(["AccountCode"]).iterrows():
        vec = np.array([float(row.get(m, 0.0)) for m in range(1, 13)], dtype=float)
        add_line(str(row["AccountCode"]), "  " + str(row["AccountName"]).replace("COGS - ", ""), vec)
    add_line("", "Total COGS", total_cogs)

    add_line("", "Gross Profit", gross_profit)

    add_blank()
    add_section("OPERATING EXPENSES (OPEX)")
    for _, row in opex.sort_values(["AccountCode"]).iterrows():
        vec = np.array([float(row.get(m, 0.0)) for m in range(1, 13)], dtype=float)
        add_line(str(row["AccountCode"]), "  " + str(row["AccountName"]).replace("Opex - ", ""), vec)
    add_line("", "Total Operating Expenses", total_opex)

    add_line("", "EBITDA", ebitda)

    add_blank()
    add_section("OTHER INCOME / (EXPENSE)")
    for _, row in other.sort_values(["AccountCode"]).iterrows():
        vec = np.array([float(row.get(m, 0.0)) for m in range(1, 13)], dtype=float)
        add_line(str(row["AccountCode"]), "  " + str(row["AccountName"]).replace("Other - ", ""), vec)
    add_line("", "Total Other", total_other)

    add_line("", "Net Income", net_income)

    df = pd.DataFrame(rows, columns=["Account", "Line Item", *MONTHS, "Total"])
    return df


def maybe_messify_cell(x: object, rng: random.Random) -> object:
    if x == "" or x is None:
        return x

    if isinstance(x, (int, float, np.floating)):
        p = rng.random()
        if p < 0.02:
            return ""
        if 0.02 <= p < 0.04:
            return "-"
        if p < 0.22:
            val = float(x)
            s = f"{abs(val):,.0f}"
            return f"({s})" if val < 0 else s
        return float(x)

    return x


# -----------------------------
# EXCEL WRITING / STYLING
# -----------------------------

def style_pnl_sheet(ws, title: str):
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True)
    total_font = Font(bold=True)

    ws.merge_cells("A1:O1")
    ws["A1"] = title
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:O2")
    ws["A2"] = f"Currency: {CURRENCY}   |   Note: values may be text-formatted / blanks to simulate real raw spreadsheets."
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:O3")
    ws["A3"] = "Use Power Query to clean, unpivot months, and build a proper fact table for Power BI."
    ws["A3"].alignment = Alignment(horizontal="center")

    for col in range(1, 16):
        c = ws.cell(row=5, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[5].height = 18

    ws.freeze_panes = "C6"

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 38
    for i in range(3, 15):  # C..N
        ws.column_dimensions[get_column_letter(i)].width = 12
    ws.column_dimensions["O"].width = 14

    max_row = ws.max_row
    for r in range(5, max_row + 1):
        for c in range(1, 16):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if c >= 3 and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0;(#,##0)'

    for r in range(6, max_row + 1):
        label = ws.cell(row=r, column=2).value
        if isinstance(label, str) and label.strip() in {
            "REVENUE", "COST OF GOODS SOLD (COGS)", "OPERATING EXPENSES (OPEX)", "OTHER INCOME / (EXPENSE)"
        }:
            for c in range(1, 16):
                ws.cell(row=r, column=c).font = section_font
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="F2F2F2")

        if isinstance(label, str) and (label.startswith("Total ") or label in {"Gross Profit", "EBITDA", "Net Income"}):
            for c in range(1, 16):
                ws.cell(row=r, column=c).font = total_font
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="FFF2CC")


def write_clean_table_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame):
    ws = wb.create_sheet(sheet_name)

    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="2F5597")
    header_font = Font(color="FFFFFF", bold=True)

    for c in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(c)].width = max(12, min(40, len(str(df.columns[c - 1])) + 2))

    for r in range(2, ws.max_row + 1):
        for c in range(1, len(df.columns) + 1):
            ws.cell(row=r, column=c).border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{ws.max_row}"


def write_messy_pnl_sheet(wb: Workbook, sheet_name: str, pnl_df: pd.DataFrame, title: str, seed: int):
    ws = wb.create_sheet(sheet_name)

    ws.append([""] * 15)  # row 4
    headers = ["Account", "Line Item", *MONTHS, "Total"]
    ws.append(headers)  # row 5

    rng = random.Random(seed)
    for _, r in pnl_df.iterrows():
        out = []
        for h in headers:
            out.append(maybe_messify_cell(r[h], rng))
        ws.append(out)

    ws.auto_filter.ref = f"A5:O{ws.max_row}"
    style_pnl_sheet(ws, title=title)


def main(output_path: str = "data/Financial_Raw_Data_2020_2024.xlsx"):
    accounts = build_accounts()
    dim_date = build_dim_date(YEARS)
    dim_account = pd.DataFrame([a.__dict__ for a in accounts])
    dim_entity = pd.DataFrame(ENTITIES)
    dim_cc = pd.DataFrame(COST_CENTERS)
    dim_product = pd.DataFrame(PRODUCTS)

    sales_ops = make_sales_operational(dim_date)
    headcount_ops = make_headcount_operational(dim_date)

    fact_pnl = build_fact_pnl(dim_date, accounts, sales_ops, headcount_ops)
    trial_balance = build_trial_balance(fact_pnl)

    pnl_wide: Dict[Tuple[int, str], pd.DataFrame] = {}
    for y in YEARS:
        for scen in ["Actual", "Budget"]:
            pnl_wide[(y, scen)] = to_pnl_lines_consolidated(fact_pnl, year=y, scenario=scen)

    wb = Workbook()
    wb.remove(wb.active)

    # README sheet inside Excel
    ws_readme = wb.create_sheet("README")
    lines = [
        f"{COMPANY} — Portfolio Dataset (2020–2024)",
        "",
        "This workbook simulates a real finance analytics engagement dataset for learning:",
        "- Messy P&L statement sheets per year (Actual/Budget) like raw client spreadsheets",
        "- Clean dimension tables (Date/Account/Entity/CostCenter/Product)",
        "- Clean fact tables (PnL Monthly, Trial Balance, Sales Ops, Headcount Ops)",
        "",
        "Suggested Power BI model:",
        "DimDate -> FactPnL_Monthly (MonthEndDate)",
        "DimAccount -> FactPnL_Monthly (AccountCode)",
        "DimEntity -> FactPnL_Monthly (Entity)",
        "DimCostCenter -> FactPnL_Monthly (CostCenter)",
        "DimProduct -> FactSales_Operational (Product)",
        "",
        "Start with the messy sheets (e.g., 'Actual P&L 2024' + 'Budget P&L 2024'):",
        "  - Clean headers/blank rows",
        "  - Unpivot months",
        "  - Add Scenario column",
        "  - Append Actual + Budget into one fact table",
    ]
    for i, line in enumerate(lines, start=1):
        ws_readme.cell(row=i, column=1, value=line)
    ws_readme.column_dimensions["A"].width = 110

    # Clean tables
    write_clean_table_sheet(wb, "DimDate", dim_date)
    write_clean_table_sheet(wb, "DimAccount", dim_account)
    write_clean_table_sheet(wb, "DimEntity", dim_entity)
    write_clean_table_sheet(wb, "DimCostCenter", dim_cc)
    write_clean_table_sheet(wb, "DimProduct", dim_product)

    write_clean_table_sheet(wb, "FactPnL_Monthly", fact_pnl)
    write_clean_table_sheet(wb, "TrialBalance_Monthly", trial_balance)
    write_clean_table_sheet(wb, "FactSales_Operational", sales_ops)
    write_clean_table_sheet(wb, "FactHeadcount", headcount_ops)

    # Messy P&L sheets per year & scenario (job-aligned)
    for y in YEARS:
        for scen in ["Actual", "Budget"]:
            sheet_name = f"{scen} P&L {y}"
            title = f"{COMPANY} — Profit & Loss (P&L) Statement — {scen} {y}"
            seed = SEED + y + (0 if scen == "Actual" else 1000)
            write_messy_pnl_sheet(wb, sheet_name, pnl_wide[(y, scen)], title, seed)

    import os
    os.makedirs("data", exist_ok=True)
    wb.save(output_path)
    print(f"✅ Created workbook: {output_path}")
    print(f"✅ Sheets: {len(wb.sheetnames)}")


if __name__ == "__main__":
    main()
